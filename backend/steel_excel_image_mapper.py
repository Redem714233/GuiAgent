from __future__ import annotations

import argparse
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _extract_timestamp_key(name: str) -> tuple[int, str]:
    match = re.search(r"_(\d{10,})\.(jpg|jpeg|png|bmp|webp)$", name.lower())
    if match:
        return int(match.group(1)), name
    return -1, name


def _ordered_images(images: list[Path], reverse_mapping: bool) -> list[Path]:
    sorted_images = sorted(images, key=lambda p: _extract_timestamp_key(p.name))
    return list(reversed(sorted_images)) if reverse_mapping else sorted_images


def _collect_images_from_dir(images_dir: Path) -> list[Path]:
    exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
    all_files = [p for p in images_dir.rglob("*") if p.is_file() and p.suffix.lower() in exts]
    if not all_files:
        raise ValueError(f"图片目录中未找到图片: {images_dir}")
    return all_files


def _extract_images_from_zip(images_zip: Path, temp_dir: Path) -> list[Path]:
    exts = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
    extracted: list[Path] = []

    with zipfile.ZipFile(images_zip, "r") as zf:
        for name in zf.namelist():
            suffix = Path(name).suffix.lower()
            if suffix not in exts:
                continue

            target_path = temp_dir / Path(name).name
            with zf.open(name, "r") as src, open(target_path, "wb") as dst:
                dst.write(src.read())
            extracted.append(target_path)

    if not extracted:
        raise ValueError(f"zip 中未找到图片: {images_zip}")

    return extracted


def embed_original_images_column(
    excel_path: Path,
    output_path: Path,
    image_paths: list[Path],
    column_name: str = "原始图片",
    image_width: int = 160,
    image_height: int = 120,
) -> tuple[int, int]:
    wb = load_workbook(excel_path)
    ws = wb.active

    data_rows = max(0, ws.max_row - 1)
    new_col_idx = ws.max_column + 1
    new_col_letter = get_column_letter(new_col_idx)

    ws.cell(row=1, column=new_col_idx, value=column_name)

    # 调整列宽（大致适配图片宽度）
    ws.column_dimensions[new_col_letter].width = max(ws.column_dimensions[new_col_letter].width or 0, 24)

    for row_offset in range(data_rows):
        row_idx = row_offset + 2
        if row_offset >= len(image_paths):
            break

        image_path = image_paths[row_offset]
        cell_ref = f"{new_col_letter}{row_idx}"

        img = XLImage(str(image_path))
        img.width = image_width
        img.height = image_height
        ws.add_image(img, cell_ref)

        # 行高单位是 points，约 px * 0.75
        target_height = max((image_height * 0.75) + 6, ws.row_dimensions[row_idx].height or 0)
        ws.row_dimensions[row_idx].height = target_height

    wb.save(output_path)
    return data_rows, len(image_paths)


def generate_excel_with_embedded_images(
    *,
    excel_path: str | Path,
    images_dir: Optional[str | Path] = None,
    images_zip: Optional[str | Path] = None,
    output_path: Optional[str | Path] = None,
    reverse_mapping: bool = True,
    column_name: str = "原始图片",
    image_width: int = 160,
    image_height: int = 120,
) -> Path:
    excel_file = Path(excel_path)
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel 不存在: {excel_file}")

    if bool(images_dir) == bool(images_zip):
        raise ValueError("必须二选一：images_dir 或 images_zip")

    if output_path:
        output_file = Path(output_path)
    else:
        output_file = excel_file.with_name(f"{excel_file.stem}_with_images{excel_file.suffix}")

    if images_dir:
        raw_images = _collect_images_from_dir(Path(images_dir))
        ordered_images = _ordered_images(raw_images, reverse_mapping=reverse_mapping)
        embed_original_images_column(
            excel_path=excel_file,
            output_path=output_file,
            image_paths=ordered_images,
            column_name=column_name,
            image_width=image_width,
            image_height=image_height,
        )
        return output_file

    with tempfile.TemporaryDirectory(prefix="steel_images_") as temp_dir:
        raw_images = _extract_images_from_zip(Path(images_zip), Path(temp_dir))
        ordered_images = _ordered_images(raw_images, reverse_mapping=reverse_mapping)
        embed_original_images_column(
            excel_path=excel_file,
            output_path=output_file,
            image_paths=ordered_images,
            column_name=column_name,
            image_width=image_width,
            image_height=image_height,
        )
    return output_file


def main() -> None:
    parser = argparse.ArgumentParser(description="为钢铁异常 Excel 增加原始图片列（真实嵌图，支持反序映射）")
    parser.add_argument("--excel", required=True, help="导出的 Excel 文件路径")
    parser.add_argument("--images-dir", help="图片目录（解压后的 picture 目录）")
    parser.add_argument("--images-zip", help="图片 zip 文件路径")
    parser.add_argument("--output", help="输出 Excel 路径（默认 *_with_images.xlsx）")
    parser.add_argument("--column-name", default="原始图片", help="新增列名")
    parser.add_argument("--no-reverse", action="store_true", help="关闭反序映射（默认开启反序）")
    parser.add_argument("--image-width", type=int, default=160, help="嵌入图片宽度（像素）")
    parser.add_argument("--image-height", type=int, default=120, help="嵌入图片高度（像素）")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 不存在: {excel_path}")

    if bool(args.images_dir) == bool(args.images_zip):
        raise ValueError("必须二选一：--images-dir 或 --images-zip")

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = excel_path.with_name(f"{excel_path.stem}_with_images{excel_path.suffix}")

    if args.images_dir:
        raw_images = _collect_images_from_dir(Path(args.images_dir))
        final_output = generate_excel_with_embedded_images(
            excel_path=excel_path,
            images_dir=args.images_dir,
            output_path=output_path,
            reverse_mapping=not args.no_reverse,
            column_name=args.column_name,
            image_width=args.image_width,
            image_height=args.image_height,
        )
        rows_count = max(0, load_workbook(excel_path).active.max_row - 1)
        images_count = len(raw_images)
    else:
        with tempfile.TemporaryDirectory(prefix="steel_images_") as temp_dir:
            raw_images = _extract_images_from_zip(Path(args.images_zip), Path(temp_dir))
            final_output = generate_excel_with_embedded_images(
                excel_path=excel_path,
                images_zip=args.images_zip,
                output_path=output_path,
                reverse_mapping=not args.no_reverse,
                column_name=args.column_name,
                image_width=args.image_width,
                image_height=args.image_height,
            )
            rows_count = max(0, load_workbook(excel_path).active.max_row - 1)
            images_count = len(raw_images)

    print(f"输入 Excel: {excel_path}")
    print(f"图片数量: {images_count}")
    print(f"数据行数: {rows_count}")
    print(f"输出 Excel: {final_output}")


if __name__ == "__main__":
    main()
